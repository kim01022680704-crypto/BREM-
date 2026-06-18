# -*- coding: utf-8 -*-
import os
import json
from openpyxl import load_workbook

base = r"C:\Users\user\Desktop\브라더 공유폴더\★BREM 경리일보★"
out = r"C:\Users\user\Desktop\BREM\scripts\revenue-xlsx-structure.json"
result = {}

for file in sorted(os.listdir(base)):
    if not file.endswith(".xlsx") or "★" not in file:
        continue
    path = os.path.join(base, file)
    wb = load_workbook(path, data_only=True, read_only=True)
    entry = {"sheets": {}}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = []
        for i, row in enumerate(ws.iter_rows(max_row=40, max_col=20, values_only=True), 1):
            cells = ["" if v is None else str(v).strip() for v in row]
            if any(cells):
                rows.append({"row": i, "cells": cells})
        entry["sheets"][name] = rows
    wb.close()
    result[file] = entry

with open(out, "w", encoding="utf-8") as f:
    json.dump(result, f, ensure_ascii=False, indent=2)
print("written", out)
