# -*- coding: utf-8 -*-
import os
from openpyxl import load_workbook

base = r"C:\Users\user\Desktop\브라더 공유폴더\★BREM 경리일보★"
for file in sorted(os.listdir(base)):
    if not file.endswith(".xlsx") or "★" not in file:
        continue
    path = os.path.join(base, file)
    print("\n========", file, "========")
    wb = load_workbook(path, data_only=True, read_only=True)
    print("Sheets:", " | ".join(wb.sheetnames))
    for name in wb.sheetnames:
        ws = wb[name]
        print(f"\n--- {name} ---")
        count = 0
        for i, row in enumerate(ws.iter_rows(max_row=35, max_col=15, values_only=True), 1):
            cells = ["" if v is None else str(v).strip() for v in row]
            if any(cells):
                print(f"{i:3}", " | ".join(cells))
                count += 1
        if count == 0:
            print("(empty)")
    wb.close()
